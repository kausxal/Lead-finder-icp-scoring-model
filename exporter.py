import csv
import os
from datetime import datetime


def export_to_csv(companies, filepath=None):
    if not companies:
        return None

    if filepath is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = f"terrascope_leads_{ts}.csv"

    fieldnames = list(companies[0].keys())
    for skip in ["score_breakdown"]:
        if skip in fieldnames:
            fieldnames.remove(skip)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for company in companies:
            row = {}
            for k in fieldnames:
                v = company.get(k)
                if isinstance(v, str):
                    row[k] = v
                elif isinstance(v, float):
                    row[k] = f"{v:.2f}"
                elif v is None:
                    row[k] = ""
                else:
                    row[k] = str(v)
            writer.writerow(row)

    return filepath


def export_google_sheets_format(companies, filepath=None):
    if not companies:
        return None

    if filepath is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = f"terrascope_leads_sheets_{ts}.csv"

    fieldnames = list(companies[0].keys())
    for skip in ["score_breakdown"]:
        if skip in fieldnames:
            fieldnames.remove(skip)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)
        for company in companies:
            row = []
            for k in fieldnames:
                v = company.get(k)
                if isinstance(v, float):
                    row.append(f"{v:.2f}")
                elif v is None:
                    row.append("")
                else:
                    row.append(str(v))
            writer.writerow(row)

    return filepath


def export_to_excel(companies, filepath=None):
    if not companies:
        return None

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return export_to_csv(companies, filepath)

    if filepath is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = f"terrascope_leads_{ts}.xlsx"

    fieldnames = list(companies[0].keys())
    for skip in ["score_breakdown"]:
        if skip in fieldnames:
            fieldnames.remove(skip)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a2035", end_color="1a2035", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Segoe UI", size=10)
    cell_align = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="2a2f3e"),
        right=Side(style="thin", color="2a2f3e"),
        top=Side(style="thin", color="2a2f3e"),
        bottom=Side(style="thin", color="2a2f3e"),
    )

    for ci, name in enumerate(fieldnames, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    for ri, company in enumerate(companies, 2):
        for ci, k in enumerate(fieldnames, 1):
            v = company.get(k)
            if isinstance(v, float):
                val = f"{v:.2f}"
            elif v is None:
                val = ""
            else:
                val = str(v)
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = cell_font
            c.alignment = cell_align
            c.border = thin_border

    ws.auto_filter.refresh = True
    wb.save(filepath)
    return filepath
